# ********************************************************************************
# Calculate Cross section properties for closed polygonal cross-sections

# Created for the Aeolus4future project
# Author: Gabriel Sabau
# ********************************************************************************
import numpy as np
from openpyxl import Workbook, load_workbook

#input Width | Height | thickness over width | thickness over height | bending radius
#input   B    |   H   |   t1  | t2    |   r
def calc_prop_RHS(B, H, t1, t2, ri):
    #The function calculates the geometrical properties of cold-formed RHS
    # A - Area of the bent arc
    # r - radius of the bent arc
    # t - thickness of the arc
    # theta - angle between 2 facets
    # ro - outside radius of the bend
    # ri - inside radius of the bend
    t = (t1 + t2) / 2
    ro = ri + t
    r = ro - t/2
    phi = 0
    theta = np.pi / 2
    A = theta * r * t
    # Ix,Iy,Ixy - second moment of area of the bend relative to center of rotation, principal axis (xc,yc)
    Ix = (ro ** 4 - ri ** 4) / 16 * (2 * theta - np.sin(2 * theta))
    Sx = (ro ** 3 - ri ** 3) * np.sin(theta) / 3
    Iy = Ix
    Sy = (ro ** 3 - ri ** 3) * (1 - np.cos(theta)) / 3
    Ixy = (ro ** 4 - ri ** 4) / 16 * (np.cos(2 * theta) - 1)

    # Ix1,Iy1 - second moment of area of the bend relative to center of rotation, axis paralel to global x,y
    Ix1 = Ix * np.cos(phi) ** 2 + Iy * np.sin(phi) ** 2 - Ixy * np.sin(2 * phi)
    Iy1 = Iy * np.cos(phi) ** 2 + Ix * np.sin(phi) ** 2 + Ixy * np.sin(2 * phi)

    # Ix0,Iy0 - second moment of area of the bend relative to origin and main axis
    yc_c = (H - 2*r - t1)/2
    xc_c = (B - 2*r - t2)/2
    ycg = Sy / A
    xcg = Sx / A
    ycg_gl = xcg * np.cos(phi) + ycg * np.sin(phi)
    xcg_gl = ycg * np.cos(phi) - xcg * np.sin(phi)
    # Ix0,Iy0 - second moment of area of the bend relative to origin and main axis
    Ixc = Ix1 + 2 * yc_c * ycg_gl * A + yc_c ** 2 * A
    Iyc = Iy1 + 2 * xc_c * xcg_gl * A + xc_c ** 2 * A
    Ac=4 * A

    #########################################
    #Plate contribution
    b = B - t2 - 2 * r
    h = H - t1 - 2 * r
    Ap = 2*b * t1 + 2 * h * t2
    # xcg,ycg - coordinates of the centroid of the plate relative to center of section (xc,yc)
    xcg = (H - t1)/2
    ycg = (B - t2)/2
    # Ix,Iy,Ixy - second moment of area of the plate relative to center of rotation, principal axis (xc,yc)
    Iyp = 2 * ((t2 ** 3) * h / 12 + b**3 * t1 / 12 + ycg**2 * h * t2)
    Ixp = 2 * ((t1 ** 3) * b / 12 + h**3 * t2 / 12 + xcg**2 * b * t1)
    Ixt = 4 * Ixc + Ixp
    Iyt = 4 * Iyc + Iyp
    At = Ac+Ap
    I = [At, min(Ixt, Iyt)]
    return I


Title = 'C:\\Users\\gabsab\\Documents\\HstrSteel\\Excel_files\\PresentData2.xlsx'  #
sheet_name = 'SHS_CF+S'
wb = load_workbook(Title, read_only=True)
sheet_position = wb.sheetnames.index(sheet_name)
wb.active = sheet_position
ws = wb.active

#extract cells containing input data
def iter_rows(wsf):
        for row in wsf['J16':'N21']:
            yield[round(cell.value, 2) for cell in row]
data_range = iter_rows(ws)
for data in data_range:
    result = calc_prop_RHS(H=data[0], B=data[1], t1=data[2], ri=data[4], t2=data[2])
    print(data)
    print(result)


